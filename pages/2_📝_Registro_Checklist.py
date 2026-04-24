import streamlit as st
import pandas as pd
import datetime
import os
import csv
import io
from openpyxl.worksheet.datavalidation import DataValidation

st.set_page_config(page_title="Registro Checklist", page_icon="📝", layout="wide")

# --- AUTENTICAÇÃO ---
if "auth_checklist" not in st.session_state:
    st.session_state.auth_checklist = False

if not st.session_state.auth_checklist:
    st.title("🔒 Acesso Restrito - Registro Checklist")
    senha_correta = os.getenv("SENHA_CHECKLIST", "Callink@02")
    senha = st.text_input("Digite a senha de acesso ao Registro:", type="password")
    if st.button("Entrar", type="primary"):
        if senha == senha_correta:
            st.session_state.auth_checklist = True
            st.rerun()
        else:
            st.error("Senha incorreta!")
    st.stop()
# ----------------------

st.title("📝 Registro de Checklist")
st.markdown("Insira os dados da vistoria abaixo. Eles serão acumulados numa prévia e você poderá enviá-los de uma vez para o sistema principal ou baixar como Excel.")

# Caminho do arquivo
CSV_FILE = "Checklist dos pisos(Planilha1).csv"

# Configuração da Blacklist de erros de português e palavras proibidas
BLACKLIST = [
    "sem lacres", # Exemplo dado pelo usuário
    "conctado", 
    "tbm", 
    "vc", 
    "pq", 
    "nao ", # para evitar 'nao' sem til
    "ta ",  # para evitar 'ta' sem acento
    "vassoura"
]

@st.cache_data(ttl=60) # Atualiza a lista a cada 60s
def get_historico_coluna(prefixo_coluna):
    try:
        try:
            df = pd.read_csv(CSV_FILE, encoding='utf-8-sig')
        except:
            if os.path.exists(CSV_FILE):
                df = pd.read_csv(CSV_FILE, encoding='latin-1')
            else:
                return []
        
        # Encontrar a coluna
        col_alvo = None
        for col in df.columns:
            if prefixo_coluna in col:
                col_alvo = col
                break
                
        if col_alvo:
            def format_number(val):
                try:
                    fval = float(val)
                    return str(int(fval)) if fval.is_integer() else str(val)
                except:
                    return str(val)
                    
            # Retorna valores únicos não vazios ordenados alfabeticamente
            opcoes = [format_number(opt) for opt in df[col_alvo].dropna().unique().tolist()]
            opcoes = sorted(list(set([str(opt).strip() for opt in opcoes if str(opt).strip() != ""])))
            return opcoes
        return []
    except Exception as e:
        return []

# Carregar opções existentes
opcoes_existentes = get_historico_coluna('Observa')
opcoes_posicao_existentes = get_historico_coluna('Posi')

# Inicializar o session_state para armazenar os registros temporários (prévia)
if 'checklist_preview' not in st.session_state:
    st.session_state.checklist_preview = []

st.markdown("---")
st.markdown("### 📱 Modo Offline (Para Múltiplos Usuários / Sem Internet)")
st.markdown("Se você vai ficar sem internet ou quer que várias pessoas usem em seus próprios celulares de forma segura, você pode baixar a planilha modelo abaixo. Preencha a planilha no celular (no Excel ou Google Sheets) e, quando tiver internet, faça o upload do arquivo salvo aqui mesmo.")

col_download, col_upload = st.columns(2)

with col_download:
    # Gerar template vazio
    df_template = pd.DataFrame(columns=["Data", "Piso", "Posição", "Módulo", "Observação"])
    buffer_template = io.BytesIO()
    with pd.ExcelWriter(buffer_template, engine='openpyxl') as writer:
        df_template.to_excel(writer, index=False, sheet_name='Checklist Offline')
        workbook = writer.book
        ws = writer.sheets['Checklist Offline']
        
        # Criar aba oculta para armazenar as opções longas de validação
        ws_listas = workbook.create_sheet('ListasOcultas')
        ws_listas.sheet_state = 'hidden'
        
        # 1. Validação de Data (A)
        dv_data = DataValidation(type="date", operator="greaterThanOrEqual", formula1="DATE(2020,1,1)")
        dv_data.promptTitle = 'Data'
        dv_data.prompt = 'Insira uma data válida'
        dv_data.errorTitle = 'Data Inválida'
        dv_data.error = 'O valor inserido deve ser uma data.'
        ws.add_data_validation(dv_data)
        dv_data.add('A2:A1048576')
        
        # 2. Validação de Piso (B)
        pisos_str = '"Piso 3,Piso 4,Piso 5,Piso 6,Niterói,UBT 4"'
        dv_piso = DataValidation(type="list", formula1=pisos_str, allow_blank=True)
        ws.add_data_validation(dv_piso)
        dv_piso.add('B2:B1048576')
        
        # 3. Validação de Posição (C) - Usando aba oculta e permitindo texto livre
        if opcoes_posicao_existentes:
            for i, opt in enumerate(opcoes_posicao_existentes, start=1):
                ws_listas.cell(row=i, column=1, value=opt)
            range_pos = f"ListasOcultas!$A$1:$A${len(opcoes_posicao_existentes)}"
            dv_pos = DataValidation(type="list", formula1=range_pos, allow_blank=True)
            dv_pos.showErrorMessage = False # Permite digitar valores que não estão na lista
            ws.add_data_validation(dv_pos)
            dv_pos.add('C2:C1048576')
            
        # 4. Validação de Módulo (D)
        modulo_str = '"' + ",".join([str(i) for i in range(1, 37)]) + '"'
        dv_mod = DataValidation(type="list", formula1=modulo_str, allow_blank=True)
        ws.add_data_validation(dv_mod)
        dv_mod.add('D2:D1048576')
        
        # 5. Validação de Observação (E) - Usando aba oculta e permitindo texto livre
        if opcoes_existentes:
            for i, opt in enumerate(opcoes_existentes, start=1):
                ws_listas.cell(row=i, column=2, value=opt)
            range_obs = f"ListasOcultas!$B$1:$B${len(opcoes_existentes)}"
            dv_obs = DataValidation(type="list", formula1=range_obs, allow_blank=True)
            dv_obs.showErrorMessage = False # Permite digitar valores que não estão na lista
            ws.add_data_validation(dv_obs)
            dv_obs.add('E2:E1048576')
        
    st.download_button(
        label="⬇️ Baixar Modelo em Branco (Excel)",
        data=buffer_template.getvalue(),
        file_name="Checklist_Template_Offline.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

with col_upload:
    with st.form("form_upload", clear_on_submit=True):
        uploaded_offline_file = st.file_uploader("⬆️ Enviar Planilha Preenchida (Offline)", type=["xlsx"])
        submit_upload = st.form_submit_button("Carregar Dados", use_container_width=True)
        
        if submit_upload and uploaded_offline_file is not None:
            try:
                df_uploaded = pd.read_excel(uploaded_offline_file)
                # Garantir formato correto das colunas
                if all(col in df_uploaded.columns for col in ["Data", "Piso", "Posição", "Observação"]):
                    if "Módulo" not in df_uploaded.columns:
                        df_uploaded["Módulo"] = ""
                    # Formatar dados corretamente
                    for index, row in df_uploaded.iterrows():
                        # Lidar com datas e evitar NaN virando "nan" string
                        if pd.isna(row['Piso']) and pd.isna(row['Posição']):
                            continue # Pula linhas inteiramente vazias do excel
                            
                        data_val = str(row['Data']) if pd.notna(row['Data']) else ""
                        if pd.notna(row['Data']) and isinstance(row['Data'], pd.Timestamp):
                            data_val = row['Data'].strftime("%d/%m/%Y")
                            
                        obs_val = str(row['Observação']) if pd.notna(row['Observação']) else ""
                        
                        # Validar Blacklist no upload offline
                        obs_invalida = False
                        for palavra in BLACKLIST:
                            if palavra.lower() in obs_val.lower():
                                st.error(f"⚠️ Linha {index+2} rejeitada: Contém palavra proibida '{palavra}' na observação.")
                                obs_invalida = True
                                break
                                
                        if not obs_invalida:
                            novo_registro = {
                                "Data": data_val,
                                "Piso": str(row['Piso']) if pd.notna(row['Piso']) else "",
                                "Posição": str(row['Posição']) if pd.notna(row['Posição']) else "",
                                "Módulo": str(row['Módulo']) if pd.notna(row['Módulo']) else "",
                                "Observação": obs_val
                            }
                            st.session_state.checklist_preview.append(novo_registro)
                    st.success("Dados válidos da planilha carregados para a prévia com sucesso!")
                else:
                    st.error("As colunas do arquivo enviado não batem com o modelo.")
            except Exception as e:
                st.error(f"Erro ao ler a planilha: {e}")

st.markdown("---")
st.markdown("### ✍️ Formulário de Entrada Direta")

# Formulário de Entrada (Online)
with st.form("form_checklist", clear_on_submit=True):
    col1, col2 = st.columns(2)
    
    with col1:
        data_input = st.date_input("Data", datetime.date.today(), format="DD/MM/YYYY")
        piso_input = st.selectbox("Piso", ["Piso 3", "Piso 4", "Piso 5", "Piso 6", "Niterói", "UBT 4"])
        col_pos, col_mod = st.columns(2)
        with col_pos:
            posicao_input = st.selectbox(
                "Posição", 
                opcoes_posicao_existentes, 
                index=None, 
                placeholder="Selecione ou digite a Posição...", 
                accept_new_options=True
            )
        with col_mod:
            opcoes_modulo = [str(i) for i in range(1, 37)]
            modulo_input = st.selectbox("Módulo", opcoes_modulo, index=None, placeholder="Selecione o módulo")
        
    with col2:
        st.markdown("**Observação**")
        obs_final = st.selectbox(
            "Selecione do histórico ou digite uma nova:", 
            opcoes_existentes, 
            index=None, 
            placeholder="Selecione ou digite a Observação...", 
            accept_new_options=True
        )
        
    submitted = st.form_submit_button("Adicionar à Prévia")
    
    if submitted:
        if not piso_input or not posicao_input or str(posicao_input).strip() == "":
            st.warning("Por favor, preencha o Piso e a Posição.")
        elif not obs_final or str(obs_final).strip() == "":
            st.warning("Por favor, digite ou selecione uma observação.")
        else:
            obs_final_str = str(obs_final).strip()
            
            # Validação contra a Blacklist (Erros de PT / Palavras proibidas)
            contem_proibida = False
            palavra_bloqueada = ""
            for palavra in BLACKLIST:
                if palavra.lower() in obs_final.lower():
                    contem_proibida = True
                    palavra_bloqueada = palavra
                    break
            
            if contem_proibida:
                st.error(f"🚫 **Erro:** A sua observação contém o termo proibido ou erro ortográfico: **'{palavra_bloqueada}'**.")
                st.info("A inserção foi bloqueada. Por favor, corrija o texto e tente novamente.")
            else:
                data_formatada = data_input.strftime("%d/%m/%Y")
                novo_registro = {
                    "Data": data_formatada,
                    "Piso": piso_input,
                    "Posição": str(posicao_input).strip(),
                    "Módulo": str(modulo_input) if modulo_input is not None else "",
                    "Observação": obs_final_str
                }
                st.session_state.checklist_preview.append(novo_registro)
                st.success("Adicionado à prévia com sucesso!")

# Exibir a prévia
st.markdown("---")
st.subheader("👀 Pré-visualização dos Dados")

if not st.session_state.checklist_preview:
    st.info("Nenhum dado na prévia ainda. Preencha o formulário acima ou faça upload de um arquivo offline.")
else:
    df_preview = pd.DataFrame(st.session_state.checklist_preview)
    st.dataframe(df_preview, use_container_width=True)
    
    st.markdown("---")
    st.markdown("### Ações")
    
    col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 1])
    
    with col_btn1:
        if st.button("Limpar Prévia", use_container_width=True):
            st.session_state.checklist_preview = []
            st.rerun()
            
    with col_btn2:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_preview.to_excel(writer, index=False, sheet_name='Checklist')
            
        st.download_button(
            label="📥 Baixar como Excel",
            data=buffer.getvalue(),
            file_name="Checklist_Preview.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
            
    with col_btn3:
        if st.button("Salvar no Sistema Central", type="primary", use_container_width=True):
            try:
                file_exists = os.path.isfile(CSV_FILE)
                with open(CSV_FILE, mode='a', newline='', encoding='utf-8-sig') as file:
                    writer = csv.writer(file)
                    if not file_exists:
                        writer.writerow(['Data', 'Piso', 'Posição', 'Módulo', 'Observação'])
                    for row in st.session_state.checklist_preview:
                        writer.writerow([row.get('Data',''), row.get('Piso',''), row.get('Posição',''), row.get('Módulo',''), row.get('Observação','')])
                
                st.session_state.checklist_preview = []
                st.success("Dados salvos com sucesso! O Dashboard já está atualizado.")
                get_historico_coluna.clear() # Limpar cache para atualizar o dropdown
                st.rerun()
                
            except Exception as e:
                st.error(f"Erro ao salvar os dados: {e}")
